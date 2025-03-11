import pandas as pd
import time
from model import chat
import numpy as np
from caculate import calculate_stats
from rouge import calculate_rouge_scores
from bleu import bleu
from openpyxl import load_workbook

def process_file(file_path, dropdown, socketio, filename, evaluation_path, tishici, use_general_algorithm):
    df = pd.read_excel(file_path, header=0)  # header=0 表示第一行为列名
    
    if '选项A' in df.columns or '模型答案' in df.columns:
        error_message = f"此为文字题，请选择文字题文件。"
        socketio.emit('error', {'message': error_message})
        raise ValueError(error_message)  # 抛出异常以停止程序

    if '问题' not in df.columns or '模型答案(文字题)' not in df.columns or '参考' not in df.columns or '标准答案(文字题)' not in df.columns or '分数' not in df.columns or '原因' not in df.columns:
        error_message = f"文件格式不正确，缺少必填列。"
        socketio.emit('error', {'message': error_message})
        raise ValueError(error_message)

    print(dropdown)
    think_box = []
    
    total_questions = len(df)
    stop_1 = 0
    for index, row in df.iterrows():
        if use_general_algorithm:
            a = pd.notna(row['标准答案(文字题)']) 
        else:
            a = True
        
        if a and pd.notna(row['问题']) and pd.notna(row['模型答案(文字题)']) and pd.isna(row['分数']) and pd.isna(row['原因']): 
            if pd.notna(row['参考']):
                reference = row['参考']
            else:
                reference = None
        
            question = row['问题']
            answer = row['模型答案(文字题)']
            time.sleep(1)
            fenshu, reason, stop, think = chat(question, reference, answer, dropdown, tishici)
            print("")             
            if stop == 1:
                stop_1 = 1
            df['原因'] = df['原因'].astype(str)  
            df.at[index, '分数'] = fenshu
            df.at[index, '原因'] = str(reason)
            if think is not None:
                think_box.append({index:think})
                
            # 更新进度
            progress = (index + 1) / total_questions * 100
            socketio.emit('progress', {'filename': filename, 'progress': progress})
        else:
            error_message = f"文件格式错误: 第 {index + 1} 行数据不正确。"    
            socketio.emit('error', {'message': error_message})
            raise ValueError(error_message)  # 抛出异常以停止程序
    
    data = df.to_dict(orient='records')
    
    if use_general_algorithm:
        socketio.emit('status', {'message': '正在计算通用模型参数......'})
        bleu_1, bleu_2, bleu_3, bleu_4 = 0, 0, 0, 0
        if total_questions < 10:
            blue_liat = list(range(total_questions))
        else:
            blue_liat = np.random.choice(range(total_questions), size=10, replace=False).tolist()
        for i in blue_liat:
            bleu_score = bleu(data[i]['标准答案(文字题)'], data[i]['模型答案(文字题)'])
            bleu_1 = bleu_1 + bleu_score[0]
            bleu_2 = bleu_2 + bleu_score[1]
            bleu_3 = bleu_3 + bleu_score[2]
            bleu_4 = bleu_4 + bleu_score[3]
            bleu_score.clear()
        
        rouge = calculate_rouge_scores(df)
    
    numbers = [item['分数'] for item in data]
    average, medium, standard_deviation = calculate_stats(numbers) 
    numbers.clear()
    
    with pd.ExcelWriter(evaluation_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='数据', index=False)
        
        workbook = writer.book
        stats_sheet = workbook['数据']
    
        if len(think_box) > 0:
            stats_sheet['M1'] = '评测think'
            for i in think_box:
                index = list(i.keys())[0]
                think = i[index]
                stats_sheet[f'M{index + 2}'] = think
            think_box.clear()
        
        stats_sheet = workbook.create_sheet(title='统计')

        # 写入统计信息到指定单元格
        stats_sheet['A1'] = '总题数'
        stats_sheet['A2'] = total_questions
        stats_sheet['B1'] = '平均分'
        stats_sheet['B2'] = average
        stats_sheet['C1'] = '中位数'
        stats_sheet['C2'] = medium
        stats_sheet['D1'] = '标准差'
        stats_sheet['D2'] = standard_deviation
        stats_sheet['A4'] = '不计算通用模型参数'
        
        if use_general_algorithm:
            stats_sheet['A4'] = 'bleu-1'
            stats_sheet['A5'] = bleu_1 / len(blue_liat)
            stats_sheet['B4'] = 'bleu-2'
            stats_sheet['B5'] = bleu_2 / len(blue_liat)
            stats_sheet['C4'] = 'bleu-3'
            stats_sheet['C5'] = bleu_3 / len(blue_liat)
            stats_sheet['D4'] = 'bleu-4'
            stats_sheet['D5'] = bleu_4 / len(blue_liat)
            rouge_sheet = workbook.create_sheet(title='rouge')
            ROUGE = pd.DataFrame(rouge).T
            ROUGE.to_excel(writer, sheet_name='rouge', index=True)

    if stop_1 == 1:  # 如果有停止标志，则发送停止信号
        socketio.emit('status', {'message': '评测中有错误，请检查结果！(-9999)'})
    else:
        socketio.emit('status', {'message': '评测成功!'})